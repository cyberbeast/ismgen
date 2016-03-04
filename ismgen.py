# STABLE BUILD
# SIGNED: CYBERBEAST

import pandas as pd 
pd.set_option('display.expand_frame_repr', False)
import math
import pprint
from collections import OrderedDict
import sys, getopt
import openpyxl
import xlsxwriter
import warnings
warnings.filterwarnings("ignore")

def restart_line():
	sys.stdout.write('\r')
	sys.stdout.flush()

def get_next_six():
	global count_host
	global count_guest
	global prev
	ret = []
	if prev is "guest":
		start = count_host
		count_host += 6
		ret = host_list[start:start+7]
		if not ret:
			return([float('nan') for i in range(7)])
		prev = "host"
	else:
		start = count_guest
		count_guest += 6
		ret = guest_list[start:start+7]
		if not ret:
			return([float('nan') for i in range(7)])
		prev = "guest"
	return(ret)

try:
	opts, args = getopt.getopt(sys.argv[1:],"h:g:m:", ["hostfile=","guestfile="])
except getopt.GetoptError:
	print('ismgen.py -h <hostfile> -g <guestfile>')
	sys.exit(2)

for opt, arg in opts:
	if opt in ("-h", "--hostfile"):
		host = pd.read_csv(arg)
	elif opt in ("-g", "--guestfile"):
		guest = pd.read_csv(arg)

host_list = host['3/4'].tolist() + host['5/6'].tolist() + host['7/8'].tolist()
guest_list = guest['3/4'].tolist() + guest['5/6'].tolist() + guest['7/8'].tolist()

prev="guest"
col_count = 1
count_host= 0
count_guest= 0

floor_list = [str(flr) + "0" for flr in range(2, 7)]
room_list = []
for floor in floor_list:
	for i in range(7):
		if i !=5:
			room_list.append(floor + str(i))

layout = OrderedDict()
layout_frames = OrderedDict()

for room in room_list:
	layout[room] = {col_num:get_next_six() for col_num in ["col" + str(num) for num in range(1,10)]}
	layout_frames[str(room)] = pd.DataFrame({k : pd.Series(v) for k, v in layout[room].items()}).fillna('----------')
	# layout_frames[str(room)].index.name = str(room)
	
pprint.pprint(layout_frames)
done_list=[]
print("\n\nGenerating Excel File! Please wait...")
wb = openpyxl.load_workbook('template.xlsx')
for room in room_list:
	restart_line()
	check_list = wb.get_sheet_names()
	if room in check_list:
		new_sheet = wb.get_sheet_by_name(room)
	else:
		new_sheet = wb.create_sheet(str(room))
	
	default_sheet = wb['Template']

	for rownum in range(7, 13):
		new_sheet.row_dimensions[rownum].height = 25

	for colname in ['B', 'C', 'D', 'F', 'G', 'H', 'J', 'K', 'L']:
		new_sheet.column_dimensions[colname].width = 18

	new_sheet.merge_cells('B1:C1')

	new_sheet.merge_cells('F3:H4')
	new_sheet.row_dimensions
	new_sheet.sheet_properties.pageSetUpPr.fitToPage = True
	new_sheet.page_setup.orientation = new_sheet.ORIENTATION_LANDSCAPE
	new_sheet.page_setup.paperSize = new_sheet.PAPERSIZE_A4
	new_sheet.print_options.horizontalCentered = True
	new_sheet.print_options.verticalCentered = True
	new_sheet.header_footer.right_header.text = str(room)
	new_sheet.header_footer.right_header.font_name = "Arial,Bold"
	new_sheet.header_footer.right_header.font_size = 125


	for row in default_sheet.rows:
		for cell in row:
			new_cell = new_sheet[str(cell.column + str(cell.row))]
			new_cell.value = cell.value
			if cell.has_style:
				new_cell.style = cell.style.copy()

	for row_t in range(7, 13):
		for indx in range(0, 6):
			new_sheet['B' + str(row_t)] = layout_frames[str(room)]['col1'].tolist()[indx]
			new_sheet['C' + str(row_t)] = layout_frames[str(room)]['col2'].tolist()[indx]
			new_sheet['D' + str(row_t)] = layout_frames[str(room)]['col3'].tolist()[indx]
			new_sheet['F' + str(row_t)] = layout_frames[str(room)]['col4'].tolist()[indx]
			new_sheet['G' + str(row_t)] = layout_frames[str(room)]['col5'].tolist()[indx]
			new_sheet['H' + str(row_t)] = layout_frames[str(room)]['col6'].tolist()[indx]
			new_sheet['J' + str(row_t)] = layout_frames[str(room)]['col7'].tolist()[indx]
			new_sheet['K' + str(row_t)] = layout_frames[str(room)]['col8'].tolist()[indx]
			new_sheet['L' + str(row_t)] = layout_frames[str(room)]['col9'].tolist()[indx]

	done_list.append(room)
	sys.stdout.write("Arranging {done_list}!".format(**locals()))
	sys.stdout.flush()

wb.save("template.xlsx")

header_logo_done_list = []
workbook = xlsxwriter.Workbook("template.xlsx")
for ws_in_wb in workbook.worksheets():
	print(ws_in_wb.get_name())
	# ws_in_wb.set_header('&L&G', {'image_left':'logo_header.png'})
	# done_list.append(ws_in_wb.get_name())
	# sys.stdout.write("Adding PESIT Logo to {header_logo_done_list}!".format(**locals()))
	# sys.stdout.flush()
workbook.close()
print("\nDone! :)")

# print(layout_frames['200']['col1'].tolist())